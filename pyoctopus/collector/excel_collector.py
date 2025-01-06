import os.path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..types import Collector, R


class CellStyle:
    def __init__(self, *,
                 font_size: int = 11,
                 font_bold=False,
                 font_color: str = '000000',
                 background_color: str = 'FFFFFF',
                 border_color: str = 'd4d4d4',
                 alignment: str = 'left',
                 delimiter: str = '\n'):
        self.font_size = font_size
        self.font_bold = font_bold
        self.font_color = font_color
        self.background_color = background_color
        self.border_color = border_color
        self.alignment = alignment
        self.delimiter = delimiter


class Column:
    def __init__(self, field: str, name: str, style: CellStyle = None):
        self.field = field
        self.name = name
        if style is None:
            self.style = CellStyle()
        else:
            self.style = style


def _set_cell_style(cell, style: CellStyle):
    cell.font = Font(size=style.font_size, bold=style.font_bold, color=style.font_color)
    cell.fill = PatternFill(fill_type="solid", start_color=style.background_color,
                            end_color=style.background_color)
    cell.alignment = Alignment(horizontal=style.alignment, vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=style.border_color),
        right=Side(style="thin", color=style.border_color),
        top=Side(style="thin", color=style.border_color),
        bottom=Side(style="thin", color=style.border_color),
    )


def new(file: str, append: bool = False, header_style: CellStyle = None,
        columns: list[Column] = None) -> Collector:
    if header_style is None:
        header_style = CellStyle(font_size=12, font_bold=True, font_color='FFFFFF', background_color='808080',
                                 alignment='left')

    if not append and os.path.exists(file):
        os.remove(file)

    if os.path.exists(file):
        wb = openpyxl.load_workbook(file)
    else:
        wb = openpyxl.Workbook()

    sheet = wb.active

    if columns and sheet.max_row == 1:
        sheet.append([c.name for c in columns])
        for cell in sheet[1]:
            _set_cell_style(cell, header_style)
        wb.save(file)

    def _collect(r: R) -> None:
        keys = r.__dict__.keys()
        _vals = []
        if not columns:
            values = [(r.__dict__.get(k, None), None) for k in keys]
        else:
            values = [(r.__dict__.get(c.field, None), c) for c in columns]
        for value in values:
            if value[0] is None:
                _vals.append('')
            elif isinstance(value[0], list):
                _vals.append((value[1].style.delimiter if value[1] else '\n').join([str(v) for v in value[0]]))
            else:
                _vals.append(str(value[0]))
        sheet.append(_vals)
        for i, cell in enumerate(sheet[sheet.max_row]):
            col_style = columns[i].style if columns and columns[i].style else CellStyle()
            _set_cell_style(cell, col_style)

        # 自动设置列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # 获取列字母
            for cell in column:
                length = sum(
                    2 if '\u4e00' <= char <= '\u9fff' else 1 for char in str(cell.value))
                max_length = max(max_length, length)
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(file)

    return _collect


def new_column(field: str, column: str, style: CellStyle = None) -> Column:
    return Column(field, column, style)


def new_cell_style(*,
                   font_size: int = 12,
                   font_bold=False,
                   font_color: str = '000000',
                   background_color: str = 'FFFFFF',
                   border_color: str = 'd4d4d4',
                   alignment: str = 'left',
                   delimiter: str = '\n') -> CellStyle:
    return CellStyle(font_size=font_size, font_bold=font_bold, font_color=font_color, border_color=border_color,
                     background_color=background_color, alignment=alignment, delimiter=delimiter)
